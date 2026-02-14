#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Construction des tables d'entrées 10–14 à partir d'un export app22

Entrée :
- Un CSV de type "submissions_export.csv" (colonnes : submission_id, payload_json, ...)
  OU un JSONL de payloads (1 JSON par ligne)

Sorties :
- 10_Repondants.csv
- 11_StatScores.csv
- 12_Genre.csv
- 13_Capacite.csv
- 14_Selections.csv

Option :
- Remplir automatiquement le classeur Classeur_analyse_app22.xlsx (écrit uniquement les feuilles 10–14).

⚠️ Si l'exécution "tourne sans fin", la cause est presque toujours l'écriture dans Excel (openpyxl)
avec une détection d'en-têtes trop large (cellules contenant des formules, ou formatage très à droite).
Cette version :
- borne la détection des en-têtes (max_cols)
- évite les boucles gigantesques
- propose un mode xlsx "fast" (écrase seulement les lignes nécessaires)

Usage :
  python build_tables_10_14_from_export_app22.py --in exports_test/submissions_export.csv --out-dir exports_test/tables
  python build_tables_10_14_from_export_app22.py --in exports_test/payloads.jsonl --out-dir exports_test/tables --fmt jsonl

  # Remplir le classeur :
  python build_tables_10_14_from_export_app22.py --in exports_test/submissions_export.csv \
    --out-xlsx Classeur_analyse_app22.xlsx --out-xlsx-filled exports_test/Classeur_analyse_app22_rempli.xlsx \
    --xlsx-mode fast
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import pandas as pd

try:
    import openpyxl
except Exception:
    openpyxl = None


# -------------------------
# Mappings (FR/EN -> colonnes canoniques du classeur)
# -------------------------

GENDER_MAP = {
    "Désagrégation par sexe": "sexe",
    "Disaggregation by sex": "sexe",
    "Sexe": "sexe",
    "Sex": "sexe",
    "Désagrégation par âge": "age",
    "Disaggregation by age": "age",
    "Âge": "age",
    "Age": "age",
    "Milieu urbain / rural": "urbain_rural",
    "Urban / rural": "urbain_rural",
    "Milieu urbain/rural": "urbain_rural",
    "Urban/rural residence": "urbain_rural",
    "Handicap": "handicap",
    "Disability": "handicap",
    "Quintile de richesse": "quintile_riche",
    "Wealth quintile": "quintile_riche",
    "Violences basées sur le genre (VBG)": "vbg",
    "Gender-based violence (GBV)": "vbg",
    "Temps domestique non rémunéré": "temps_domestique",
    "Unpaid domestic work": "temps_domestique",
}

CAPACITY_MAP = {
    "Compétences statistiques disponibles": "comp_stats",
    "Available statistical skills": "comp_stats",
    "Accès aux données administratives": "acces_data_admin",
    "Access to administrative data": "acces_data_admin",
    "Financement disponible": "financement",
    "Available funding": "financement",
    "Financement": "financement",
    "Funding": "financement",
    "Outils numériques (collecte, traitement, diffusion)": "outils_numeriques",
    "Digital tools (collection, processing, dissemination)": "outils_numeriques",
    "Outils numériques": "outils_numeriques",
    "Digital tools": "outils_numeriques",
    "Cadre juridique pour le partage de données": "cadre_juridique",
    "Legal framework for data sharing": "cadre_juridique",
    "Cadre juridique": "cadre_juridique",
    "Legal framework": "cadre_juridique",
    "Coordination interinstitutionnelle": "coordination",
    "Inter-institutional coordination": "coordination",
    "Coordination institutionnelle": "coordination",
    "Institutional coordination": "coordination",
}


def normalize_availability(v_raw: Any, scoring_version: Any) -> int:
    """
    Compatibilité : si vieux exports (availability inversée), normaliser sur 'Bonne=3'.
    Ici, si scoring_version est absent ou < 3, on inverse (1<->3).
    """
    try:
        iv = int(v_raw)
    except Exception:
        return 0
    if iv not in (0, 1, 2, 3):
        return 0
    try:
        sv = int(scoring_version)
    except Exception:
        sv = 0
    if sv >= 3:
        return iv
    if iv == 1:
        return 3
    if iv == 3:
        return 1
    return iv


def read_payloads(in_path: Path, fmt: str) -> List[Dict[str, Any]]:
    payloads: List[Dict[str, Any]] = []
    if fmt == "csv":
        df = pd.read_csv(in_path, dtype=str).fillna("")
        if "payload_json" not in df.columns:
            raise ValueError("CSV attendu : colonne payload_json manquante")
        for _, r in df.iterrows():
            try:
                payloads.append(json.loads(r["payload_json"]))
            except Exception:
                continue
    else:
        with open(in_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    payloads.append(json.loads(line))
                except Exception:
                    continue
    return payloads


def build_tables(payloads: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows10 = []
    rows11 = []
    rows12 = []
    rows13 = []
    rows14 = []

    for p in payloads:
        sid = str(p.get("submission_id", "")).strip()
        if not sid:
            continue

        # -------- 10_Repondants --------
        top5 = p.get("top5_domains", [])
        if not isinstance(top5, list):
            top5 = []
        top5 = top5 + [""] * 5

        rows10.append({
            "submission_id": sid,
            "submitted_at": p.get("submitted_at_utc", ""),
            "email": p.get("email", ""),
            "organisation": p.get("organisation", ""),
            "pays_iso3": p.get("pays", ""),
            "type_acteur": p.get("type_acteur", ""),
            "fonction": p.get("fonction", ""),
            "scope": p.get("scope", ""),
            "snds_status": p.get("snds_status", ""),
            "top_domain_1": top5[0],
            "top_domain_2": top5[1],
            "top_domain_3": top5[2],
            "top_domain_4": top5[3],
            "top_domain_5": top5[4],
        })

        # -------- 11_StatScores --------
        selected_by_domain = p.get("selected_by_domain", {})
        scoring = p.get("scoring", {})
        scoring_version = p.get("scoring_version", p.get("scoringVersion", ""))

        if isinstance(selected_by_domain, dict):
            for dom, stats in selected_by_domain.items():
                if not isinstance(stats, list):
                    continue
                for st in stats:
                    st = str(st).strip()
                    if not st:
                        continue
                    sc = scoring.get(st, {}) if isinstance(scoring, dict) else {}
                    d = sc.get("demand", "")
                    a_raw = sc.get("availability", sc.get("gap", 0))
                    f = sc.get("feasibility", "")
                    a = normalize_availability(a_raw, scoring_version)
                    try:
                        d = int(d)
                    except Exception:
                        d = ""
                    try:
                        f = int(f)
                    except Exception:
                        f = ""
                    rows11.append({
                        "submission_id": sid,
                        "domain_code": str(dom).strip(),
                        "stat_code": st,
                        "demand_score": d,
                        "availability_score": a,
                        "feasibility_score": f,
                    })

        # -------- 12_Genre --------
        g = p.get("gender_table", {})
        out_g = {"submission_id": sid, "sexe": "", "age": "", "urbain_rural": "", "handicap": "", "quintile_riche": "", "vbg": "", "temps_domestique": ""}
        if isinstance(g, dict):
            for k, v in g.items():
                canon = GENDER_MAP.get(str(k).strip(), None)
                if canon is not None:
                    out_g[canon] = str(v).strip()
        rows12.append(out_g)

        # -------- 13_Capacite --------
        c = p.get("capacity_table", {})
        out_c = {"submission_id": sid, "comp_stats": "", "acces_data_admin": "", "financement": "", "outils_numeriques": "", "cadre_juridique": "", "coordination": ""}
        if isinstance(c, dict):
            for k, v in c.items():
                canon = CAPACITY_MAP.get(str(k).strip(), None)
                if canon is not None:
                    out_c[canon] = str(v).strip()
        rows13.append(out_c)

        # -------- 14_Selections (long) --------
        q_list = p.get("quality_expectations", [])
        q_other = str(p.get("quality_other", "")).strip()
        if isinstance(q_list, list):
            for opt in q_list:
                opt_s = str(opt).strip()
                if not opt_s:
                    continue
                detail = q_other if ("Autre" in opt_s) or ("Other" in opt_s) else ""
                rows14.append({"submission_id": sid, "category": "QUALITE", "option": opt_s, "option_detail": detail})

        d_list = p.get("dissemination_channels", [])
        d_other = str(p.get("dissemination_other", "")).strip()
        if isinstance(d_list, list):
            for opt in d_list:
                opt_s = str(opt).strip()
                if not opt_s:
                    continue
                detail = d_other if ("Autre" in opt_s) or ("Other" in opt_s) else ""
                rows14.append({"submission_id": sid, "category": "DIFFUSION", "option": opt_s, "option_detail": detail})

        s_list = p.get("data_sources", [])
        s_other = str(p.get("data_sources_other", "")).strip()
        if isinstance(s_list, list):
            for opt in s_list:
                opt_s = str(opt).strip()
                if not opt_s:
                    continue
                detail = s_other if ("Autres" in opt_s) or ("Other" in opt_s) else ""
                rows14.append({"submission_id": sid, "category": "SOURCE", "option": opt_s, "option_detail": detail})

    return (
        pd.DataFrame(rows10),
        pd.DataFrame(rows11),
        pd.DataFrame(rows12),
        pd.DataFrame(rows13),
        pd.DataFrame(rows14),
    )


def write_tables_csv(out_dir: Path, tables: Tuple[pd.DataFrame, ...]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    names = ["10_Repondants.csv", "11_StatScores.csv", "12_Genre.csv", "13_Capacite.csv", "14_Selections.csv"]
    for name, df in zip(names, tables):
        df.to_csv(out_dir / name, index=False, encoding="utf-8-sig")


_FORMULA_RE = re.compile(r"^\s*=")

def _get_headers(ws, header_row: int, max_cols: int) -> List[str]:
    """
    Détection robuste : les en-têtes doivent être des valeurs texte "simples".
    On ignore les cellules contenant des formules (commençant par '=') pour éviter de parcourir 10 000 colonnes.
    """
    headers: List[str] = []
    for col in range(1, max_cols + 1):
        v = ws.cell(header_row, col).value
        if v is None:
            if headers:
                break
            continue
        s = str(v).strip()
        if s == "":
            if headers:
                break
            continue
        if _FORMULA_RE.match(s):
            # un en-tête ne devrait pas être une formule : on stoppe dès que les en-têtes ont commencé
            if headers:
                break
            # si formule au début (rare) on continue quelques colonnes
            continue
        headers.append(s)
    return headers


def write_to_workbook(template_xlsx: Path, out_xlsx: Path, tables: Tuple[pd.DataFrame, ...], header_row: int, max_cols: int, xlsx_mode: str) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl n'est pas installé dans cet environnement.")
    wb = openpyxl.load_workbook(template_xlsx)

    mapping = {
        "10_Repondants": tables[0],
        "11_StatScores": tables[1],
        "12_Genre": tables[2],
        "13_Capacite": tables[3],
        "14_Selections": tables[4],
    }

    for sheet, df in mapping.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        headers = _get_headers(ws, header_row=header_row, max_cols=max_cols)
        if not headers:
            # fallback : utiliser les colonnes df si pas d'en-têtes dans le modèle
            headers = list(df.columns)

        start_row = header_row + 1

        # Mode d'écriture
        # - fast : on écrase les len(df)+50 lignes, c'est rapide et évite les boucles longues
        # - clean : on efface un peu plus (len(df)+500) pour éviter les résidus
        extra_clear = 50 if xlsx_mode == "fast" else 500
        clear_rows = max(len(df) + extra_clear, 200)

        # Effacement borné
        for r in range(start_row, start_row + clear_rows):
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).value = None

        # Écriture
        for i, (_, row) in enumerate(df.iterrows()):
            r = start_row + i
            for c, h in enumerate(headers, start=1):
                ws.cell(r, c).value = row.get(h, "")

    wb.save(out_xlsx)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", required=True, help="Chemin du fichier export (CSV ou JSONL)")
    ap.add_argument("--fmt", choices=["csv", "jsonl"], default="csv", help="Format d'entrée")
    ap.add_argument("--out-dir", default="tables_10_14", help="Dossier de sortie (CSVs)")

    # Excel
    ap.add_argument("--out-xlsx", default="", help="Classeur modèle (ex : Classeur_analyse_app22.xlsx)")
    ap.add_argument("--out-xlsx-filled", default="", help="Classeur rempli en sortie (ex : Classeur_analyse_app22_rempli.xlsx)")
    ap.add_argument("--header-row", type=int, default=4, help="Ligne des en-têtes dans les feuilles 10–14 (défaut : 4)")
    ap.add_argument("--max-cols", type=int, default=80, help="Nombre maximum de colonnes à scanner pour détecter les en-têtes (défaut : 80)")
    ap.add_argument("--xlsx-mode", choices=["fast", "clean"], default="fast", help="fast (rapide) ou clean (efface plus)")

    args = ap.parse_args()

    in_path = Path(args.in_path)
    payloads = read_payloads(in_path, fmt=args.fmt)
    tables = build_tables(payloads)

    out_dir = Path(args.out_dir)
    write_tables_csv(out_dir, tables)

    if args.out_xlsx and args.out_xlsx_filled:
        write_to_workbook(
            template_xlsx=Path(args.out_xlsx),
            out_xlsx=Path(args.out_xlsx_filled),
            tables=tables,
            header_row=int(args.header_row),
            max_cols=int(args.max_cols),
            xlsx_mode=str(args.xlsx_mode),
        )

    print("OK : tables 10–14 générées")
    print(f"- Dossier : {out_dir}")
    if args.out_xlsx and args.out_xlsx_filled:
        print(f"- Classeur : {args.out_xlsx_filled}")


if __name__ == "__main__":
    main()
